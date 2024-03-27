import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
from openpyxl import Workbook, load_workbook
import re
from openpyxl.utils import get_column_letter


def read_urls_from_csv(csv_file_path):
    df = pd.read_csv(csv_file_path)
    urls = df["Object_URL"].tolist()
    return urls


def scrape_address_latitude_longitude(url):
    try:
        headers_request = {'Accept-Language': 'en-US,en;q=0.9'}
        response = requests.get(url, headers=headers_request)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            address = soup.find('h1').text.strip()
            script_tags = soup.find_all('script')
            latitude = None
            longitude = None
            for script in script_tags:
                if 'initMap' in script.text:
                    match = re.search(r'initMap\(([\d.-]+\.\d+),\s*([\d.-]+\.\d+)', script.text)
                    if match:
                        latitude = match.group(1)
                        longitude = match.group(2)
                        break
            return address, latitude, longitude
        else:
            return None, None, None
    except Exception as e:
        return None, None, None


def write_to_excel(data, excel_file_path):
    if os.path.exists(excel_file_path):
        wb = load_workbook(excel_file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Address', 'Latitude', 'Longitude'])  # Add column headers if new workbook

    for row in data:
        ws.append([row.get('Address'), row.get('Latitude'), row.get('Longitude')])

    wb.save(excel_file_path)


def chunk_urls(urls, chunk_size):
    for i in range(0, len(urls), chunk_size):
        yield urls[i:i + chunk_size]


input_csv_path = 'Building_data.csv'
output_excel_path = 'location12deci.xlsx'  # Output Excel file

urls = read_urls_from_csv(input_csv_path)

for index, chunk in enumerate(chunk_urls(urls, 800), start=1):
    scraped_data = []
    for url in chunk:
        address, latitude, longitude = scrape_address_latitude_longitude(url)
        scraped_data.append({'Address': address or 'Not Found', 'Latitude': latitude or 'Not Found',
                             'Longitude': longitude or 'Not Found'})

    write_to_excel(scraped_data, output_excel_path)
    processed_count = index * 800 if index * 800 < len(urls) else len(urls)
    print(f"Processed {processed_count} URLs.")

    if processed_count < len(urls):
        print("Waiting 6 minutes before processing the next chunk...")
        time.sleep(360)

print("All data has been processed and written to the Excel file.")
